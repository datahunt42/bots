from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from aiogram.types import ContentType, PreCheckoutQuery, LabeledPrice, ReplyKeyboardRemove, ReplyKeyboardMarkup, \
    KeyboardButton
from openpyxl import load_workbook

# хэндлеры
said_handler = ['1', '2', '3', '4', '5', '6', 'Выбрать заново']
detail_handler = ['Кран', 'Липучка', 'Кнопка', 'Крутилка', 'Джойстик', 'Спинер']
# основные переменные
i = 0
a = 0
API_TOKEN = '5473096213:AAFIMTv7Q5ELnFXAZ0uj6iougXZgVLE7tJ8'
photo = 'https://chudo-udo.info/media/k2/items/cache/b98eb4e6d4e5af022817653939abd5f0_XL.jpg'
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)
fn = 'Telegram.xlsx'
prices = [LabeledPrice('Игрушка', 10000)]
all_details = ['/Кран', '/Липучка', '/Кнопка', '/Крутилка', '/Джойстик', '/Спинер']
# кнопки выбора грани
keyboard_sides = types.ReplyKeyboardMarkup(resize_keyboard=True)
buttons_sides = ['/1', '/2', '/3', '/4', '/5', '/6', ]
keyboard_sides.add(*buttons_sides)
# кнопки выбора детали
keyboard_details = types.ReplyKeyboardMarkup(resize_keyboard=True)
buttons_details = all_details
keyboard_details.add(*buttons_details)
# Кнопка начала действий(Заказа)
keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
button = ['/Сделать_заказ']
keyboard.add(*button)  #
# Кнопки оплаты
buy_keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
button3 = ['/Купить', 'Выбрать заново']
buy_keyboard.row(*button3)


@dp.message_handler(commands="start")
async def intro_function(message: types.Message):
    await message.answer("Добро пожаловать! Выберите действие: ", reply_markup=keyboard)


@dp.message_handler(commands=['Сделать_заказ'])
async def selecting(message: types.Message):
    await message.answer('Ваше имя для составления заказа', reply_markup=ReplyKeyboardRemove())

    @dp.message_handler()
    async def first_name(message: types.Message):
        if message.text:
            print(message.text)
            wb = load_workbook(fn)
            wb.create_sheet(message.text)
            wb.save(fn)
            wb.close()
            await bot.send_photo(chat_id=message.chat.id, photo=photo, caption='Выберите грань',
                                 reply_markup=keyboard_sides)


@dp.message_handler(commands=said_handler)
async def grains(message: types.Message):
    global i
    wb = load_workbook(fn)
    await bot.send_message(chat_id=message.chat.id, text='Выберите деталь', reply_markup=keyboard_details)
    i += 1
    ws1 = wb.active
    ws1['A' + str(i)] = message.text
    wb.save(fn)
    wb.close()
    print(i)


@dp.message_handler(commands=detail_handler)
async def restrn(message: types.Message):
    global a
    global i
    wb = load_workbook(fn)
    await bot.send_photo(chat_id=message.chat.id, photo=photo, caption='Выберите грань', reply_markup=keyboard_sides)
    a += 1
    ws1 = wb.active
    ws1['B' + str(i)] = message.text
    wb.save(fn)
    wb.close()
    print(a)
    if a == 6:
        await message.delete()
        await bot.send_message(message.chat.id, text='Если все выбрано верно, подтвердите оплату',
                               reply_markup=buy_keyboard)
        i = 0
        a = 0

    @dp.message_handler()
    async def tryagain(message: types.Message):
        if message.text == 'Выбрать заново':
            await bot.send_photo(chat_id=message.chat.id, photo=photo, caption='Выберите грань',
                                 reply_markup=keyboard_sides)


@dp.message_handler(commands='Купить')
async def payment(message: types.Message):
    await bot.send_invoice(chat_id=message.chat.id,
                           title='Игрушка',
                           description='четотам',
                           payload='al-use',
                           provider_token='1744374395:TEST:a83375fc3b7a41d1a47d',
                           prices=prices,
                           currency='RUB'

                           )


@dp.pre_checkout_query_handler(lambda q: True)
async def checkout_query(pre_checkout_query: PreCheckoutQuery):
    await bot.answer_pre_checkout_query(pre_checkout_query.id, ok=True)


@dp.message_handler(content_types=ContentType.SUCCESSFUL_PAYMENT)
async def successful_payment(message: types.Message):
    doc = open('Telegram.xlsx', 'rb')
    await bot.send_message(message.chat.id, text="оплата прошла успешно", reply_markup=keyboard)
    await bot.send_document(chat_id='896595378', document=doc)


if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
